import os
import threading
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from ..models import PREvent, ReviewResponse

logger_lock = threading.Lock()

class ExcelLogger:
    """Logs PR reviews and updates execution status in a shared Excel spreadsheet."""

    def __init__(self, excel_path: str) -> None:
        """
        Initialize ExcelLogger with path to Excel file.

        Args:
            excel_path: Local filesystem path where Excel logs are saved.
                        Can be a directory (appends default filename) or a full .xlsx path.
        """
        # If path is a directory or doesn't end with .xlsx/.xls, treat it as a directory
        if not excel_path.endswith(('.xlsx', '.xls')):
            excel_path = os.path.join(excel_path, "codelens_reviews.xlsx")
        self.excel_path = excel_path
        # Ensure parent directory exists
        excel_dir = os.path.dirname(self.excel_path)
        if excel_dir:
            os.makedirs(excel_dir, exist_ok=True)

    def _init_workbook(self) -> None:
        """
        Initialize the Excel workbook with headers if it does not already exist.
        """
        if not os.path.exists(self.excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PR Reviews"
            headers = [
                "date", "pr_id", "repo", "author", "jira_ticket", "approval", "confidence",
                "critical_issues", "high_issues", "total_issues", "summary", "tokens_used", "latency_ms", "status"
            ]
            ws.append(headers)
            wb.save(self.excel_path)
            wb.close()

    def log_review(self, pr_event: PREvent, review: ReviewResponse) -> None:
        """
        Append a PR review record to the Excel worksheet.

        Args:
            pr_event: The Pull Request event.
            review: Final review response.
        """
        with logger_lock:
            self._init_workbook()

            # Count issues by severity
            critical_cnt = sum(1 for i in review.issues if str(i.severity).upper() == "CRITICAL")
            high_cnt = sum(1 for i in review.issues if str(i.severity).upper() == "HIGH")
            total_cnt = len(review.issues)

            row = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                pr_event.pr_id,
                pr_event.repo,
                pr_event.author,
                pr_event.jira_ticket_id or "",
                str(review.approval),
                review.confidence,
                critical_cnt,
                high_cnt,
                total_cnt,
                review.summary,
                review.tokens_used,
                review.latency_ms,
                "Reviewed"
            ]

            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb["PR Reviews"]
            ws.append(row)

            # Auto-size columns to fit values neatly
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

            wb.save(self.excel_path)
            wb.close()

    def get_history(self, repo: str = None) -> pd.DataFrame:
        """
        Retrieve PR review logging history.

        Args:
            repo: Optional repository filter (e.g. 'owner/repo').

        Returns:
            A pandas DataFrame containing logs.
        """
        if not os.path.exists(self.excel_path):
            return pd.DataFrame()

        with logger_lock:
            try:
                df = pd.read_excel(self.excel_path, sheet_name="PR Reviews")
                if repo:
                    df = df[df["repo"] == repo]
                return df
            except Exception:
                return pd.DataFrame()

    def update_status(self, pr_id: str, status: str) -> None:
        """
        Update the status column for a specific PR ID record.

        Args:
            pr_id: PR ID number to update.
            status: New status message.
        """
        if not os.path.exists(self.excel_path):
            return

        with logger_lock:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb["PR Reviews"]

            headers = [cell.value for cell in ws[1]]
            try:
                pr_id_col = headers.index("pr_id") + 1
                status_col = headers.index("status") + 1
            except ValueError:
                wb.close()
                return

            for row in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=pr_id_col).value
                if cell_val is not None and str(cell_val).strip() == str(pr_id).strip():
                    ws.cell(row=row, column=status_col, value=status)

            wb.save(self.excel_path)
            wb.close()
